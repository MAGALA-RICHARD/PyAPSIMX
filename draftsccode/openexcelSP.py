import time
import logging
import arcpy
import openpyxl
import os
import traceback
import arcpy
class ReadExcel:
      def __init__(self, path, excelname):
        '''
        costructs class for creating lookup info
        paramters
        ------------
        
        excelname: name with the land use and their corresponding nesting weights
        
        parameters
        '''
        fullpath = os.path.join(path, excelname)
        if (os.path.isfile(fullpath)) and (fullpath.endswith('.xlsx')):
            self.ExcelName =  excelname
            self.path  = path
            self.fullpath = fullpath
        else:
          logging.error("File does not exists make sure it is the specified in the working directory and no csv file is allowed")
      
      def ExtractInfo(self):
        '''
        
        '''
        try:
            start = time.perf_counter()
            self.objectid = []
            self.lon =[]
            self.lat  = []
              #load the excel book
            book = openpyxl.load_workbook(self.fullpath)
          # get the sheet names 
            sheet = book.sheetnames
            for index, sheets in enumerate(sheet):
             if sheets =="Lookup_info":
               lookupinfo = book.worksheets[index]
               for i in range(1, lookupinfo.max_row+1):
                    if i !=1:
                      self.objectid.append(lookupinfo.cell(row = i, column = 1).value)
                      self.lon.append(lookupinfo.cell(row =i, column = 2).value)
                      self.lat.append(lookupinfo.cell(row =i, column = 3).value)  
            if "Lookup_info" not in sheet:
              print("No look up information sheet found")
            if sheets =="management":
               management = book.worksheets[index]
            else:
              print("No management sheet found")
          # iterate through ground nesting sheet
          
            return self
        except:
           tb = sys.exc_info()[2]
           tbinfo = traceback.format_tb(tb)[0]
      
           # Concatenate information together concerning the error into a message string
           pymsg = "PYTHON ERRORS:\nTraceback info:\n" + tbinfo + "\nError Info:\n" + str(sys.exc_info()[1])
           print(pymsg + "\n")
      
           if arcpy.GetMessages(2) not in pymsg:
              msgs = "ArcPy ERRORS:\n" + arcpy.GetMessages(2) + "\n"
              arcpy.AddError(msgs)
              print(msgs)
      
        finally:
            end = time.perf_counter()
            print(f'reading simulation info: {end-start} seconds')
class ReadExcel2:
      def __init__(self, fullpath):
        '''
        costructs class for creating lookup info. readexcel2 takes in a full path for the excelname
        
        paramters
        ------------
        
        excelname: name with the simulation info
        
        parameters
        '''
        self.fullpath = fullpath
        if (os.path.isfile(self.fullpath)) and (self.fullpath.endswith('.xlsx')):
            pass
        else:
          logging.error("File does not exists make sure it is the specified in the working directory and no csv file is allowed")
      
      def ExtractInfo(self):
        '''
        
        '''
        try:
            start = time.perf_counter()
            self.objectid = []
            self.lonlat =[]
            self.lat  = []
              #load the excel book
            book = openpyxl.load_workbook(self.fullpath)
          # get the sheet names 
            sheet = book.sheetnames
            for index, sheets in enumerate(sheet):
             if sheets =="Lookup_info":
               lookupinfo = book.worksheets[index]
               for i in range(1, lookupinfo.max_row+1):
                    if i !=1:
                      self.objectid.append(lookupinfo.cell(row = i, column = 1).value)
                      self.lonlat.append([lookupinfo.cell(row =i, column = 2).value, lookupinfo.cell(row =i, column = 3).value])
                      self.lat.append(lookupinfo.cell(row =i, column = 3).value)  
            if "Lookup_info" not in sheet:
              print("No look up information sheet found")
            if sheets =="management":
               management = book.worksheets[index]
            else:
              print("No management sheet found took")
          # iterate through ground nesting sheet
          
            return self
        except:
           tb = sys.exc_info()[2]
           tbinfo = traceback.format_tb(tb)[0]
      
           # Concatenate information together concerning the error into a message string
           pymsg = "PYTHON ERRORS:\nTraceback info:\n" + tbinfo + "\nError Info:\n" + str(sys.exc_info()[1])
           print(pymsg + "\n")
      
        finally:
            end = time.perf_counter()
            print(f'reading simulation info: {end-start} seconds')

